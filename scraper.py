"""
Scraper de Hiring Room con interfaz gráfica.

El Excel de entrada debe tener columnas:
  - 'empresa' (o 'empresas', 'company', 'nombre')
  - 'enlace'  (o 'url', 'link', 'enlaces')
  - 'sector'  (opcional - si está, permite filtrar qué sectores procesar)

Genera Excel con:
  empresa | puesto | descripcion | requisitos | publicado | enlace

Creado por Pablo Margewka.

Uso:
    python cv_gui.py
"""

import os
import re
import asyncio
import threading
import subprocess
import sys
import unicodedata
from datetime import datetime
from pathlib import Path

import pandas as pd
import tkinter as tk
from tkinter import ttk, filedialog, messagebox, scrolledtext

from playwright.async_api import async_playwright

COL_EMPRESA = ["empresa", "empresas", "company", "nombre", "nombre_empresa"]
COL_ENLACE = ["enlace", "enlaces", "url", "link", "links", "portal"]
COL_SECTOR = ["sector", "sectores", "rubro", "rubros", "industria", "categoria"]

TIMEOUT = 60000
CONCURRENCIA = 5
WAIT_POST_LOAD = 1500
POLL_INTERVAL = 500
POLL_MAX = 60

COLUMNAS_OUTPUT = [
    "empresa", "puesto", "descripcion", "requisitos", "publicado", "enlace"
]

# === PALETA DE COLORES - AZUL CLARO ===
COLOR_BG = "#eef8ff"
COLOR_PANEL = "#ffffff"
COLOR_PANEL_SUAVE = "#f8fcff"
COLOR_PRIMARIO = "#38bdf8"      # azul claro
COLOR_PRIMARIO_HOVER = "#0ea5e9"
COLOR_ACENTO = "#22c55e"        # verde
COLOR_TEXTO = "#0f172a"
COLOR_TEXTO_SUAVE = "#64748b"
COLOR_BORDE = "#bfdbfe"
COLOR_PELIGRO = "#ef4444"
COLOR_WARNING = "#f59e0b"
COLOR_SIDEBAR = "#075985"


# ============================================================
# FILTROS PREVIOS AL SCRAPING / NORMALIZACIÓN
# ============================================================

def normalizar_texto(texto):
    """Convierte texto a minúsculas, sin acentos y con espacios normalizados."""
    if not texto:
        return ""
    texto = str(texto).lower()
    texto = unicodedata.normalize("NFD", texto)
    texto = texto.encode("ascii", "ignore").decode("utf-8")
    texto = re.sub(r"\s+", " ", texto)
    return texto.strip()


def publicado_a_dias(publicado):
    """
    Convierte el texto relativo de HiringRoom a días aproximados.
    Ejemplos: Hoy=0, Ayer=1, Hace 2 semanas=14, Hace 1 mes=30.
    """
    if not publicado:
        return None

    p = normalizar_texto(publicado)

    if "recien" in p or "hoy" in p:
        return 0
    if "ayer" in p:
        return 1

    m = re.search(r"hace\s+(\d+)\s+(\w+)", p)
    if not m:
        return None

    cantidad = int(m.group(1))
    unidad = m.group(2)

    if unidad.startswith("minuto") or unidad.startswith("hora"):
        return 0
    if unidad.startswith("dia"):
        return cantidad
    if unidad.startswith("semana"):
        return cantidad * 7
    if unidad.startswith("mes"):
        return cantidad * 30
    if unidad.startswith("ano"):
        return cantidad * 365

    return None


def cumple_filtro_tiempo(publicado, filtro_tiempo):
    """Filtro previo por antigüedad de publicación."""
    if filtro_tiempo == "Todas":
        return True

    dias = publicado_a_dias(publicado)

    # Si no se pudo detectar fecha y el usuario pidió fecha específica, se descarta.
    if dias is None:
        return False

    if filtro_tiempo == "Hoy":
        return dias == 0
    if filtro_tiempo == "Semana":
        return dias <= 7
    if filtro_tiempo == "Mes":
        return dias <= 30

    return True


def cumple_filtro_keywords(oferta, keywords):
    """
    Filtro por palabras clave con lógica OR.

    Mejora importante:
    - No distingue mayúsculas/minúsculas.
    - No distingue acentos.
    - Evita falsos positivos por substring.
      Ejemplo: la keyword "it" ya NO matchea partes sueltas dentro de otras palabras.
    - Permite escribir keywords separadas por espacios o comas.
    - Agrega variantes útiles para búsquedas IT.
    """
    keywords = normalizar_texto(keywords)

    if not keywords:
        return True

    texto_oferta = normalizar_texto(
        f"{oferta.get('empresa', '')} "
        f"{oferta.get('puesto', '')} "
        f"{oferta.get('descripcion', '')} "
        f"{oferta.get('requisitos', '')}"
    )

    # separa por comas, punto y coma o espacios
    palabras = [p.strip() for p in re.split(r"[,;\s]+", keywords) if p.strip()]

    # Expansiones para no depender de una sola forma de escribir la búsqueda
    expansiones = {
        "pasantia": ["pasantia", "pasantias", "pasante", "pasantes", "trainee", "internship"],
        "pasantias": ["pasantia", "pasantias", "pasante", "pasantes", "trainee", "internship"],
        "pasante": ["pasantia", "pasantias", "pasante", "pasantes", "trainee", "internship"],
        "sistema": ["sistema", "sistemas"],
        "sistemas": ["sistema", "sistemas"],
        "informatica": ["informatica", "informatico", "informaticos", "computacion", "tecnologia"],
        "it": ["it", "i.t", "i.t.", "tecnologia", "tecnologias", "sistemas", "informatica"],
        "soporte": ["soporte", "helpdesk", "help desk", "mesa de ayuda", "service desk"],
    }

    patrones = []

    for palabra in palabras:
        variantes = expansiones.get(palabra, [palabra])

        for variante in variantes:
            variante = normalizar_texto(variante)

            if not variante:
                continue

            # Frases como "mesa de ayuda" se buscan como frase completa.
            if " " in variante:
                patrones.append(re.escape(variante))
            else:
                # Palabras completas: evita que "it" matchee dentro de otras palabras.
                patrones.append(rf"\b{re.escape(variante)}\b")

    return any(re.search(patron, texto_oferta) for patron in patrones)


# ============================================================
# CONTROL DE PAUSA / DETENCIÓN
# ============================================================

class ControladorEjecucion:
    def __init__(self):
        self.pausado = False
        self.detenido = False
        self._evento_pausa = asyncio.Event()
        self._evento_pausa.set()

    def pausar(self):
        self.pausado = True
        self._evento_pausa.clear()

    def reanudar(self):
        self.pausado = False
        self._evento_pausa.set()

    def detener(self):
        self.detenido = True
        self._evento_pausa.set()

    async def esperar_si_pausado(self):
        await self._evento_pausa.wait()


# ============================================================
# EXTRACCIÓN DE FECHA RELATIVA
# ============================================================

PATRON_FECHA = re.compile(
    r"(Hace\s+\d+\s+(?:día|dias|días|semana|semanas|mes|meses|año|años|hora|horas|minuto|minutos)|Hoy|Ayer|Recién publicada)",
    re.IGNORECASE
)


def extraer_publicado(texto):
    if not texto:
        return ""
    match = PATRON_FECHA.search(texto)
    if match:
        return match.group(1).strip()
    return ""


# ============================================================
# LIMPIEZA Y PARSEO DE TEXTO
# ============================================================

def limpiar_texto(texto):
    if not texto:
        return ""
    basura = [
        r"Filtrar por:.*?Borrar todo",
        r"Ver\s*\d+\s*vacantes?",
        r"Referentes\s*Preguntas frecuentes",
        r"Career site",
        r"Subir CV a base general",
        r"Compartir\s*Preguntas frecuentes",
        r"Powered by.*?HiringRoom",
        r"Hiring Room es una plataforma.*?contratación de la empresa\.",
        r"¿Cómo me postulo\?.*?Recomendaciones",
        r"Contacto con soporte",
        r"Tu consulta fue enviada exitosamente",
        r"Aquí podrás escribirnos.*?responderá en\s*breve\.",
        r"Selecciona un motivo.*?Otros",
        r"Nombre \*.*?Comentarios \*",
        r"Cancelar\s*Volver\s*Enviar",
        r"Facebook\s*LinkedIn\s*Twitter\s*Whatsapp",
        r"Información de contacto",
        r"\d+-\d+ de \d+ vacantes?",
        r"No hemos encontrado vacantes.*?búsqueda",
        r"No hay vacantes publicadas",
        r"Todos los filtros",
        r"Volver al listado",
        r"Postularme",
        r"Postular",
        r"Compartir oferta",
    ]
    for patron in basura:
        texto = re.sub(patron, "", texto, flags=re.IGNORECASE | re.DOTALL)
    texto = re.sub(r"\n\s*\n\s*\n+", "\n\n", texto)
    texto = re.sub(r"[ \t]+", " ", texto)
    return texto.strip()


def split_descripcion_requisitos(texto):
    if not texto:
        return "", ""
    patrones = [
        r"\n\s*Requisitos[:\s]*\n",
        r"\n\s*Requerimientos[:\s]*\n",
        r"\n\s*¿Qué buscamos\??\s*\n",
        r"\n\s*Qué buscamos\??\s*\n",
        r"\n\s*Buscamos[:\s]*\n",
        r"\n\s*Perfil buscado[:\s]*\n",
        r"\n\s*Perfil[:\s]*\n",
        r"\n\s*Conocimientos requeridos[:\s]*\n",
        r"\n\s*Conocimientos[:\s]*\n",
        r"\n\s*Necesitamos que[:\s]*",
        r"\n\s*Es importante que[:\s]*",
        r"\n\s*Tus principales tareas serán[:\s]*",
        r"\n\s*Lo que necesitás[:\s]*",
        r"\n\s*Te pedimos[:\s]*",
    ]
    for patron in patrones:
        match = re.search(patron, texto, re.IGNORECASE)
        if match:
            return texto[:match.start()].strip(), texto[match.start():].strip()
    return texto.strip(), ""


# ============================================================
# SCRAPING
# ============================================================

async def get_job_links_con_fechas(page, portal_url, log_fn, controlador, filtro_tiempo="Todas"):
    log_fn(f"  → Cargando portal: {portal_url}")
    try:
        await page.goto(portal_url, wait_until="domcontentloaded", timeout=TIMEOUT)
    except Exception as e:
        log_fn(f"  ✗ Error cargando portal: {e}")
        return {}

    log_fn("  ⏳ Esperando vacantes...")
    enlaces = []
    for _ in range(POLL_MAX):
        if controlador.detenido:
            return {}
        await page.wait_for_timeout(POLL_INTERVAL)
        try:
            enlaces = await page.eval_on_selector_all(
                "a[href*='get_vacancy']",
                "els => [...new Set(els.map(e => e.href))]"
            )
            if enlaces:
                break
        except Exception:
            continue

    fechas_por_url = {}
    descartadas_fecha = 0

    for url in enlaces:
        try:
            texto_card = await page.evaluate(
                """(url) => {
                    const links = document.querySelectorAll('a[href*="get_vacancy"]');
                    for (const link of links) {
                        if (link.href === url) {
                            return link.innerText || link.textContent || '';
                        }
                    }
                    return '';
                }""",
                url
            )
            publicado = extraer_publicado(texto_card)
        except Exception:
            publicado = ""

        # FILTRO PREVIO REAL: si la fecha del listado no cumple, ni siquiera se abre el detalle.
        if not cumple_filtro_tiempo(publicado, filtro_tiempo):
            descartadas_fecha += 1
            continue

        fechas_por_url[url] = publicado

    log_fn(f"  ✓ {len(enlaces)} vacantes encontradas en este portal")
    if filtro_tiempo != "Todas":
        log_fn(f"  📅 {len(fechas_por_url)} pasan filtro de fecha ({filtro_tiempo}); {descartadas_fecha} descartadas")

    return fechas_por_url


async def parse_job_detail(context, url, empresa, fecha_listado, log_fn,
                           semaforo, controlador, resultados, filtro_keywords=""):
    if controlador.detenido:
        return

    async with semaforo:
        await controlador.esperar_si_pausado()
        if controlador.detenido:
            return

        page = await context.new_page()
        try:
            await page.goto(url, wait_until="domcontentloaded", timeout=TIMEOUT)
            await page.wait_for_timeout(WAIT_POST_LOAD)

            titulo = None
            for selector in ["h1", "h2", ".vacancy-title", ".job-title"]:
                try:
                    el = await page.query_selector(selector)
                    if el:
                        txt = (await el.inner_text()).strip()
                        if txt and len(txt) > 3:
                            titulo = txt
                            break
                except Exception:
                    continue

            texto_crudo = ""
            for selector in [".vacancy-description", ".job-description",
                             "[class*='description']", "main", "article"]:
                try:
                    el = await page.query_selector(selector)
                    if el:
                        txt = (await el.inner_text()).strip()
                        if len(txt) > 100:
                            texto_crudo = txt
                            break
                except Exception:
                    continue

            if not texto_crudo:
                try:
                    body = await page.query_selector("body")
                    if body:
                        texto_crudo = (await body.inner_text()).strip()
                except Exception:
                    pass

            publicado = fecha_listado or extraer_publicado(texto_crudo)
            texto_limpio = limpiar_texto(texto_crudo)
            descripcion, requisitos = split_descripcion_requisitos(texto_limpio)

            resultado = {
                "empresa": empresa,
                "puesto": titulo or "Sin título",
                "descripcion": descripcion,
                "requisitos": requisitos,
                "publicado": publicado,
                "enlace": url,
            }
            # FILTRO DE KEYWORDS ANTES DE GUARDAR LA OFERTA
            # Se evalúa sobre título, descripción y requisitos ya extraídos.
            if not cumple_filtro_keywords(resultado, filtro_keywords):
                log_fn(f"        🔎 Descartada por keywords: {resultado['puesto'][:55]}")
                return

            resultados.append(resultado)

            fecha_log = f" [{publicado}]" if publicado else ""
            log_fn(f"        ✓ {resultado['puesto'][:55]}{fecha_log}")

        except Exception as e:
            log_fn(f"        ✗ Error en {url}: {e}")
            resultados.append({
                "empresa": empresa,
                "puesto": "Error",
                "descripcion": str(e),
                "requisitos": "",
                "publicado": "",
                "enlace": url,
            })
        finally:
            await page.close()


async def scrape_portales(portales, headless, log_fn, progress_fn, controlador, resultados, filtro_tiempo="Todas", filtro_keywords=""):
    async with async_playwright() as p:
        browser = await p.chromium.launch(headless=headless)
        context = await browser.new_context(
            user_agent="Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 "
                       "(KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36",
            locale="es-AR",
        )
        page_listado = await context.new_page()
        semaforo = asyncio.Semaphore(CONCURRENCIA)

        try:
            for idx_portal, (empresa, portal_url) in enumerate(portales, 1):
                if controlador.detenido:
                    log_fn("\n⏹ Detenido por el usuario.")
                    break
                await controlador.esperar_si_pausado()
                if controlador.detenido:
                    break

                log_fn(f"\n[{idx_portal}/{len(portales)}] {empresa}")
                fechas_por_url = await get_job_links_con_fechas(
                    page_listado, portal_url, log_fn, controlador, filtro_tiempo
                )

                if not fechas_por_url:
                    progress_fn(idx_portal, len(portales))
                    continue

                log_fn(f"  ⚡ Procesando {len(fechas_por_url)} vacantes en paralelo "
                       f"(hasta {CONCURRENCIA} a la vez)...")

                tareas = [
                    parse_job_detail(context, url, empresa, fecha,
                                     log_fn, semaforo, controlador, resultados, filtro_keywords)
                    for url, fecha in fechas_por_url.items()
                ]
                await asyncio.gather(*tareas, return_exceptions=True)
                progress_fn(idx_portal, len(portales))
        finally:
            await browser.close()


def guardar_excel(jobs, ruta_salida):
    df = pd.DataFrame(jobs, columns=COLUMNAS_OUTPUT)

    with pd.ExcelWriter(ruta_salida, engine="openpyxl") as writer:
        df.to_excel(writer, index=False, sheet_name="Ofertas")
        ws = writer.sheets["Ofertas"]

        anchos = {"A": 22, "B": 38, "C": 65, "D": 65, "E": 18, "F": 50}
        for col, ancho in anchos.items():
            ws.column_dimensions[col].width = ancho

        from openpyxl.styles import Alignment, Font, PatternFill

        for row in ws.iter_rows(min_row=2):
            for cell in row:
                cell.alignment = Alignment(wrap_text=True, vertical="top")

        for fila_num in range(2, len(df) + 2):
            ws.row_dimensions[fila_num].height = 15

        header_fill = PatternFill(start_color="2563eb", end_color="2563eb", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[1].height = 22

        ws.freeze_panes = "A2"


# ============================================================
# LECTURA DEL EXCEL DE ENTRADA
# ============================================================

def detectar_columna(df, candidatos):
    cols_lower = {c.lower().strip(): c for c in df.columns}
    for cand in candidatos:
        if cand in cols_lower:
            return cols_lower[cand]
    return None


def cargar_portales_desde_excel(path):
    """
    Devuelve lista de dicts: [{empresa, enlace, sector}, ...]
    Si la columna sector no existe, todos quedan con sector="Sin clasificar".
    """
    df = pd.read_excel(path)
    col_empresa = detectar_columna(df, COL_EMPRESA)
    col_enlace = detectar_columna(df, COL_ENLACE)
    col_sector = detectar_columna(df, COL_SECTOR)

    if not col_empresa:
        raise ValueError(
            f"No encontré una columna de empresa.\n"
            f"Esperaba alguna de: {', '.join(COL_EMPRESA)}\n"
            f"Encontré: {', '.join(df.columns)}"
        )
    if not col_enlace:
        raise ValueError(
            f"No encontré una columna de enlace.\n"
            f"Esperaba alguna de: {', '.join(COL_ENLACE)}\n"
            f"Encontré: {', '.join(df.columns)}"
        )

    portales = []
    for _, row in df.iterrows():
        empresa = str(row[col_empresa]).strip() if pd.notna(row[col_empresa]) else ""
        url = str(row[col_enlace]).strip() if pd.notna(row[col_enlace]) else ""
        if col_sector and pd.notna(row[col_sector]):
            sector = str(row[col_sector]).strip()
        else:
            sector = "Sin clasificar"

        if empresa and url and url.lower().startswith("http"):
            portales.append({"empresa": empresa, "enlace": url, "sector": sector})
    return portales, bool(col_sector)


# ============================================================
# INTERFAZ GRÁFICA
# ============================================================


# ============================================================
# INTERFAZ GRÁFICA - VERSIÓN CON BARRA LATERAL Y SCROLL
# ============================================================

class ScrollableFrame(ttk.Frame):
    """Frame con scroll vertical para que la UI nunca quede cortada."""
    def __init__(self, parent, bg=COLOR_BG):
        super().__init__(parent)
        self.canvas = tk.Canvas(parent, bg=bg, highlightthickness=0, bd=0)
        self.scrollbar = ttk.Scrollbar(parent, orient="vertical", command=self.canvas.yview)
        self.scrollable = ttk.Frame(self.canvas, style="Main.TFrame")

        self.window_id = self.canvas.create_window((0, 0), window=self.scrollable, anchor="nw")
        self.canvas.configure(yscrollcommand=self.scrollbar.set)

        self.canvas.pack(side="left", fill="both", expand=True)
        self.scrollbar.pack(side="right", fill="y")

        self.scrollable.bind("<Configure>", self._on_frame_configure)
        self.canvas.bind("<Configure>", self._on_canvas_configure)
        self.canvas.bind_all("<MouseWheel>", self._on_mousewheel)

    def _on_frame_configure(self, _event=None):
        self.canvas.configure(scrollregion=self.canvas.bbox("all"))

    def _on_canvas_configure(self, event):
        self.canvas.itemconfig(self.window_id, width=event.width)

    def _on_mousewheel(self, event):
        try:
            self.canvas.yview_scroll(int(-1 * (event.delta / 120)), "units")
        except tk.TclError:
            pass


class App:
    def __init__(self, root):
        self.root = root
        self.root.title("Scraper Hiring Room - Creado por Pablo Margewka")
        self.root.geometry("1180x760")
        self.root.minsize(980, 640)
        self.root.configure(bg=COLOR_BG)

        self.archivo_entrada = None
        self.portales_todos = []
        self.tiene_columna_sector = False
        self.sectores_vars = {}
        self.sectores_widgets = {}
        self.sector_panel_visible = False
        self.archivo_salida = None
        self.scraping_en_curso = False
        self.t_inicio = None

        self.resultados = []
        self.controlador = None
        self.loop_scraping = None

        # Filtros previos al scraping
        self.var_filtro_tiempo = tk.StringVar(value="Todas")
        self.var_keywords = tk.StringVar(value="")
        self.var_export_dir = tk.StringVar(value="")

        self._configurar_estilo()
        self._build_ui()

    def _configurar_estilo(self):
        style = ttk.Style()
        style.theme_use("clam")
        style.configure(".", font=("Segoe UI", 10))

        style.configure("Main.TFrame", background=COLOR_BG)
        style.configure("Sidebar.TFrame", background=COLOR_SIDEBAR)
        style.configure("SidebarInner.TFrame", background=COLOR_SIDEBAR)
        style.configure("Card.TFrame", background=COLOR_PANEL)
        style.configure("Row.TFrame", background=COLOR_PANEL)

        style.configure("TLabel", background=COLOR_BG, foreground=COLOR_TEXTO)
        style.configure("Card.TLabel", background=COLOR_PANEL, foreground=COLOR_TEXTO)
        style.configure("Hint.TLabel", background=COLOR_PANEL, foreground=COLOR_TEXTO_SUAVE, font=("Segoe UI", 9))
        style.configure("Title.TLabel", background=COLOR_BG, foreground=COLOR_TEXTO, font=("Segoe UI", 18, "bold"))
        style.configure("Subtitle.TLabel", background=COLOR_BG, foreground=COLOR_TEXTO_SUAVE, font=("Segoe UI", 10))
        style.configure("SidebarTitle.TLabel", background=COLOR_SIDEBAR, foreground="white", font=("Segoe UI", 16, "bold"))
        style.configure("SidebarText.TLabel", background=COLOR_SIDEBAR, foreground="#e0f2fe", font=("Segoe UI", 9))
        style.configure("SidebarMetric.TLabel", background=COLOR_SIDEBAR, foreground="white", font=("Segoe UI", 18, "bold"))
        style.configure("SidebarSmall.TLabel", background=COLOR_SIDEBAR, foreground="#bae6fd", font=("Segoe UI", 9))
        style.configure("SectionTitle.TLabel", background=COLOR_PANEL, foreground=COLOR_TEXTO, font=("Segoe UI", 12, "bold"))
        style.configure("Footer.TLabel", background=COLOR_BG, foreground=COLOR_TEXTO_SUAVE, font=("Segoe UI", 9, "italic"))

        style.configure("Card.TLabelframe", background=COLOR_PANEL, borderwidth=1, relief="solid", bordercolor=COLOR_BORDE)
        style.configure("Card.TLabelframe.Label", background=COLOR_PANEL, foreground=COLOR_PRIMARIO, font=("Segoe UI", 10, "bold"))

        style.configure("Primary.TButton", background=COLOR_PRIMARIO, foreground="white", font=("Segoe UI", 10, "bold"), padding=(14, 9), borderwidth=0)
        style.map("Primary.TButton", background=[("active", COLOR_PRIMARIO_HOVER), ("disabled", "#9ca3af")])
        style.configure("SidebarPrimary.TButton", background=COLOR_PRIMARIO, foreground="white", font=("Segoe UI", 10, "bold"), padding=(12, 10), borderwidth=0)
        style.map("SidebarPrimary.TButton", background=[("active", COLOR_PRIMARIO_HOVER), ("disabled", "#374151")], foreground=[("disabled", "#9ca3af")])
        style.configure("Secondary.TButton", background="white", foreground=COLOR_TEXTO, font=("Segoe UI", 10), padding=(12, 8), borderwidth=1, relief="solid")
        style.map("Secondary.TButton", background=[("active", "#f3f4f6"), ("disabled", "#f9fafb")], foreground=[("disabled", "#9ca3af")])
        style.configure("Danger.TButton", background="white", foreground=COLOR_PELIGRO, font=("Segoe UI", 10), padding=(12, 8), borderwidth=1, relief="solid")
        style.map("Danger.TButton", background=[("active", "#fee2e2"), ("disabled", "#f9fafb")], foreground=[("disabled", "#9ca3af")])

        style.configure("Treeview", background="white", fieldbackground="white", foreground=COLOR_TEXTO, borderwidth=1, relief="solid", rowheight=28)
        style.configure("Treeview.Heading", background="#dbeafe", foreground=COLOR_TEXTO, font=("Segoe UI", 9, "bold"), relief="flat")
        style.configure("TProgressbar", background=COLOR_PRIMARIO, troughcolor=COLOR_BORDE, borderwidth=0, thickness=12)
        style.configure("TCheckbutton", background=COLOR_PANEL, foreground=COLOR_TEXTO, font=("Segoe UI", 10))
        style.configure("Sidebar.TCheckbutton", background=COLOR_SIDEBAR, foreground="#e0f2fe", font=("Segoe UI", 10))

    def _card(self, parent, title, subtitle=None):
        card = ttk.LabelFrame(parent, text=f" {title} ", style="Card.TLabelframe", padding=14)
        card.pack(fill="x", pady=(0, 12))
        if subtitle:
            ttk.Label(card, text=subtitle, style="Hint.TLabel").pack(anchor="w", pady=(0, 10))
        return card

    def _build_ui(self):
        shell = ttk.Frame(self.root, style="Main.TFrame")
        shell.pack(fill="both", expand=True)

        self.sidebar = ttk.Frame(shell, style="Sidebar.TFrame", width=245)
        self.sidebar.pack(side="left", fill="y")
        self.sidebar.pack_propagate(False)

        side = ttk.Frame(self.sidebar, style="SidebarInner.TFrame", padding=(18, 18))
        side.pack(fill="both", expand=True)

        ttk.Label(side, text="💼 Hiring Room", style="SidebarTitle.TLabel").pack(anchor="w")
        ttk.Label(side, text="Scraper de ofertas laborales", style="SidebarText.TLabel").pack(anchor="w", pady=(2, 18))

        self.side_estado = ttk.Label(side, text="Listo", style="SidebarMetric.TLabel")
        self.side_estado.pack(anchor="w")
        ttk.Label(side, text="Estado actual", style="SidebarSmall.TLabel").pack(anchor="w", pady=(0, 18))

        self.side_portales = ttk.Label(side, text="0", style="SidebarMetric.TLabel")
        self.side_portales.pack(anchor="w")
        ttk.Label(side, text="Portales seleccionados", style="SidebarSmall.TLabel").pack(anchor="w", pady=(0, 18))

        self.side_ofertas = ttk.Label(side, text="0", style="SidebarMetric.TLabel")
        self.side_ofertas.pack(anchor="w")
        ttk.Label(side, text="Ofertas extraídas", style="SidebarSmall.TLabel").pack(anchor="w", pady=(0, 22))

        ttk.Separator(side, orient="horizontal").pack(fill="x", pady=(0, 14))

        self.btn_procesar = ttk.Button(side, text="▶  Iniciar", command=self.iniciar_scraping, style="SidebarPrimary.TButton", state="disabled")
        self.btn_procesar.pack(fill="x", pady=(0, 8))
        self.btn_pausar = ttk.Button(side, text="⏸  Pausar", command=self.toggle_pausa, style="Secondary.TButton", state="disabled")
        self.btn_pausar.pack(fill="x", pady=(0, 8))
        self.btn_exportar_y_pausar = ttk.Button(side, text="💾⏸  Exportar y pausar", command=self.exportar_y_pausar, style="Secondary.TButton", state="disabled")
        self.btn_exportar_y_pausar.pack(fill="x", pady=(0, 8))
        self.btn_detener = ttk.Button(side, text="⏹  Detener", command=self.detener_scraping, style="Danger.TButton", state="disabled")
        self.btn_detener.pack(fill="x", pady=(0, 8))
        self.btn_exportar_parcial = ttk.Button(side, text="💾  Exportar parcial", command=self.exportar_parcial, style="Secondary.TButton", state="disabled")
        self.btn_exportar_parcial.pack(fill="x", pady=(0, 14))

        self.var_visible = tk.BooleanVar(value=False)
        ttk.Checkbutton(side, text="Mostrar navegador", variable=self.var_visible, style="Sidebar.TCheckbutton").pack(anchor="w", pady=(2, 0))

        ttk.Label(side, text="Creado por\nPablo Margewka", style="SidebarSmall.TLabel", justify="left").pack(side="bottom", anchor="w")

        content_wrap = ttk.Frame(shell, style="Main.TFrame")
        content_wrap.pack(side="left", fill="both", expand=True)

        scroll_area = ScrollableFrame(content_wrap, bg=COLOR_BG)
        self.main = scroll_area.scrollable

        header = ttk.Frame(self.main, style="Main.TFrame", padding=(22, 18, 22, 6))
        header.pack(fill="x")
        ttk.Label(header, text="📊 Scraper Hiring Room", style="Title.TLabel").pack(anchor="w")
        ttk.Label(header, text="Carga un Excel, filtrá por sectores, procesá portales y exportá resultados parciales o finales.", style="Subtitle.TLabel").pack(anchor="w", pady=(2, 0))

        body = ttk.Frame(self.main, style="Main.TFrame", padding=(22, 8, 22, 18))
        body.pack(fill="both", expand=True)

        sec1 = self._card(body, "1. Archivo de entrada", "El Excel debe tener columnas 'empresa', 'enlace' y opcionalmente 'sector'.")
        row1 = ttk.Frame(sec1, style="Row.TFrame")
        row1.pack(fill="x")
        self.btn_cargar = ttk.Button(row1, text="📁  Seleccionar Excel", command=self.cargar_archivo, style="Secondary.TButton")
        self.btn_cargar.pack(side="left")
        self.lbl_archivo = ttk.Label(row1, text="Ningún archivo cargado", style="Hint.TLabel")
        self.lbl_archivo.pack(side="left", padx=12)

        self.sec_sectores = self._card(body, "2. Filtro por sector compacto")
        top_sec = ttk.Frame(self.sec_sectores, style="Row.TFrame")
        top_sec.pack(fill="x")
        self.lbl_sectores_hint = ttk.Label(top_sec, text="Cargá un archivo para ver los sectores disponibles.", style="Hint.TLabel")
        self.lbl_sectores_hint.pack(side="left", anchor="w", fill="x", expand=True)
        self.btn_toggle_sectores = ttk.Button(top_sec, text="Mostrar sectores", command=self.toggle_panel_sectores, style="Secondary.TButton", state="disabled")
        self.btn_toggle_sectores.pack(side="right")

        self.panel_sectores = ttk.Frame(self.sec_sectores, style="Row.TFrame")
        self.var_buscar_sector = tk.StringVar()
        self.var_buscar_sector.trace_add("write", lambda *_: self._filtrar_checkboxes_sectores())

        search_row = ttk.Frame(self.panel_sectores, style="Row.TFrame")
        search_row.pack(fill="x", pady=(10, 8))
        ttk.Label(search_row, text="Buscar sector:", style="Card.TLabel").pack(side="left")
        self.entry_buscar_sector = ttk.Entry(search_row, textvariable=self.var_buscar_sector)
        self.entry_buscar_sector.pack(side="left", fill="x", expand=True, padx=(8, 10))

        self.botones_sectores = ttk.Frame(self.panel_sectores, style="Row.TFrame")
        self.botones_sectores.pack(fill="x", pady=(0, 8))

        list_wrap = ttk.Frame(self.panel_sectores, style="Row.TFrame")
        list_wrap.pack(fill="x")
        self.frame_sectores_canvas = tk.Canvas(list_wrap, height=96, bg=COLOR_PANEL, highlightthickness=1, highlightbackground=COLOR_BORDE)
        self.frame_sectores_scroll = ttk.Scrollbar(list_wrap, orient="vertical", command=self.frame_sectores_canvas.yview)
        self.frame_sectores = ttk.Frame(self.frame_sectores_canvas, style="Row.TFrame")
        self.frame_sectores_window = self.frame_sectores_canvas.create_window((0, 0), window=self.frame_sectores, anchor="nw")
        self.frame_sectores_canvas.configure(yscrollcommand=self.frame_sectores_scroll.set)
        self.frame_sectores_canvas.pack(side="left", fill="x", expand=True)
        self.frame_sectores_scroll.pack(side="right", fill="y")
        self.frame_sectores.bind("<Configure>", lambda e: self.frame_sectores_canvas.configure(scrollregion=self.frame_sectores_canvas.bbox("all")))
        self.frame_sectores_canvas.bind("<Configure>", lambda e: self.frame_sectores_canvas.itemconfig(self.frame_sectores_window, width=e.width))

        sec_filtros_previos = self._card(
            body,
            "3. 🎯 Filtros antes del scraping",
            "Estos filtros se aplican antes de procesar/guardar ofertas. La fecha descarta vacantes desde el listado; las keywords descartan ofertas antes de guardarlas."
        )

        filtros_row = ttk.Frame(sec_filtros_previos, style="Row.TFrame")
        filtros_row.pack(fill="x", pady=(0, 8))

        ttk.Label(filtros_row, text="📅 Publicadas:", style="Card.TLabel").pack(side="left")
        self.combo_filtro_tiempo = ttk.Combobox(
            filtros_row,
            textvariable=self.var_filtro_tiempo,
            values=["Todas", "Hoy", "Semana", "Mes"],
            state="readonly",
            width=12
        )
        self.combo_filtro_tiempo.pack(side="left", padx=(8, 18))

        ttk.Label(filtros_row, text="🔑 Palabras clave:", style="Card.TLabel").pack(side="left")
        self.entry_keywords = ttk.Entry(filtros_row, textvariable=self.var_keywords, width=42)
        self.entry_keywords.pack(side="left", fill="x", expand=True, padx=(8, 10))

        ttk.Button(
            filtros_row,
            text="🧹 Limpiar",
            command=self.limpiar_filtros_previos,
            style="Secondary.TButton"
        ).pack(side="left")

        export_row = ttk.Frame(sec_filtros_previos, style="Row.TFrame")
        export_row.pack(fill="x", pady=(6, 0))

        ttk.Label(export_row, text="📁 Exportar en:", style="Card.TLabel").pack(side="left")
        self.lbl_export_dir = ttk.Label(export_row, textvariable=self.var_export_dir, style="Hint.TLabel")
        self.lbl_export_dir.pack(side="left", fill="x", expand=True, padx=(8, 10))
        ttk.Button(
            export_row,
            text="Elegir carpeta",
            command=self.seleccionar_carpeta_exportacion,
            style="Secondary.TButton"
        ).pack(side="left")

        sec3 = self._card(body, "4. 🌐 Portales a procesar")
        self.lbl_count = ttk.Label(sec3, text="0 portales seleccionados", style="Hint.TLabel")
        self.lbl_count.pack(anchor="w", pady=(0, 8))
        self.var_buscar_portal = tk.StringVar()
        self.var_buscar_portal.trace_add("write", lambda *_: self._refrescar_lista_portales())
        portal_search_row = ttk.Frame(sec3, style="Row.TFrame")
        portal_search_row.pack(fill="x", pady=(0, 8))
        ttk.Label(portal_search_row, text="Buscar empresa/URL:", style="Card.TLabel").pack(side="left")
        ttk.Entry(portal_search_row, textvariable=self.var_buscar_portal).pack(side="left", fill="x", expand=True, padx=(8, 0))
        tree_container = ttk.Frame(sec3, style="Row.TFrame")
        tree_container.pack(fill="both", expand=True)
        cols = ("empresa", "sector", "enlace")
        self.tree = ttk.Treeview(tree_container, columns=cols, show="headings", height=8)
        self.tree.heading("empresa", text="Empresa")
        self.tree.heading("sector", text="Sector")
        self.tree.heading("enlace", text="Enlace")
        self.tree.column("empresa", width=190, minwidth=130, stretch=False)
        self.tree.column("sector", width=160, minwidth=100, stretch=False)
        self.tree.column("enlace", width=560, minwidth=300, stretch=True)
        self.tree.pack(side="left", fill="both", expand=True)
        scroll_y = ttk.Scrollbar(tree_container, orient="vertical", command=self.tree.yview)
        scroll_y.pack(side="right", fill="y")
        self.tree.configure(yscrollcommand=scroll_y.set)

        sec4 = self._card(body, "5. 📈 Progreso")
        prog_row = ttk.Frame(sec4, style="Row.TFrame")
        prog_row.pack(fill="x")
        self.progress = ttk.Progressbar(prog_row, mode="determinate", style="TProgressbar")
        self.progress.pack(side="left", fill="x", expand=True)
        self.lbl_progreso = ttk.Label(prog_row, text="", style="Card.TLabel")
        self.lbl_progreso.pack(side="left", padx=10)
        self.lbl_estado = ttk.Label(prog_row, text="Listo", style="Card.TLabel", foreground=COLOR_PRIMARIO)
        self.lbl_estado.pack(side="left", padx=10)

        sec5 = self._card(body, "6. 📝 Log de actividad")
        self.log = scrolledtext.ScrolledText(sec5, height=12, wrap="word", font=("Consolas", 9), background="#082f49", foreground="#e0f2fe", insertbackground="white", relief="flat", borderwidth=0)
        self.log.pack(fill="both", expand=True)
        self.log.configure(state="disabled")

    def _set_estado_visual(self, texto, color=COLOR_PRIMARIO):
        self.lbl_estado.config(text=texto, foreground=color)
        self.side_estado.config(text=texto.replace("●", "").replace("✓", "").replace("⏸", "").replace("⏹", "").strip() or texto)

    def log_msg(self, msg):
        def _append():
            self.log.configure(state="normal")
            self.log.insert("end", msg + "\n")
            self.log.see("end")
            self.log.configure(state="disabled")
            count = len(self.resultados)
            self.side_ofertas.config(text=str(count))
            self.btn_exportar_parcial.config(text=f"💾  Exportar parcial ({count})" if count else "💾  Exportar parcial")
        self.root.after(0, _append)

    def actualizar_progreso(self, actual, total):
        def _upd():
            self.progress["maximum"] = total
            self.progress["value"] = actual
            self.lbl_progreso.config(text=f"{actual} / {total} portales")
        self.root.after(0, _upd)

    def limpiar_filtros_previos(self):
        self.var_filtro_tiempo.set("Todas")
        self.var_keywords.set("")

    def seleccionar_carpeta_exportacion(self):
        carpeta = filedialog.askdirectory(title="Seleccionar carpeta de exportación")
        if carpeta:
            self.var_export_dir.set(carpeta)
            self.log_msg(f"📁 Carpeta de exportación seleccionada: {carpeta}")

    def _carpeta_exportacion(self):
        carpeta = self.var_export_dir.get().strip()
        if carpeta:
            return Path(carpeta)
        if self.archivo_entrada:
            return Path(self.archivo_entrada).parent
        return Path.cwd()

    def cargar_archivo(self):
        path = filedialog.askopenfilename(title="Seleccionar Excel de entrada", filetypes=[("Excel", "*.xlsx *.xls"), ("Todos", "*.*")])
        if not path:
            return
        try:
            portales, tiene_sector = cargar_portales_desde_excel(path)
        except Exception as e:
            messagebox.showerror("Error al leer Excel", str(e))
            return
        if not portales:
            messagebox.showwarning("Sin datos", "El Excel no tiene filas válidas con empresa y enlace.")
            return

        self.archivo_entrada = path
        self.var_export_dir.set(str(Path(path).parent))
        self.portales_todos = portales
        self.tiene_columna_sector = tiene_sector
        self.lbl_archivo.config(text=Path(path).name, foreground=COLOR_TEXTO)
        self._construir_filtro_sectores()
        self._refrescar_lista_portales()
        self.log_msg(f"✓ Cargados {len(portales)} portales desde {Path(path).name}")
        if tiene_sector:
            sectores_unicos = sorted({p['sector'] for p in portales})
            self.log_msg(f"  Sectores detectados: {', '.join(sectores_unicos)}")

    def _construir_filtro_sectores(self):
        for widget in self.frame_sectores.winfo_children():
            widget.destroy()
        for widget in self.botones_sectores.winfo_children():
            widget.destroy()
        self.sectores_vars.clear()
        self.sectores_widgets.clear()
        self.var_buscar_sector.set("")
        self.panel_sectores.pack_forget()
        self.sector_panel_visible = False
        self.btn_toggle_sectores.config(text="Mostrar sectores", state="disabled")

        sectores = sorted({p["sector"] for p in self.portales_todos})
        if not self.tiene_columna_sector:
            self.lbl_sectores_hint.config(text="ℹ El Excel no tiene columna 'sector'. Se procesarán todos los portales.")
            return

        self.btn_toggle_sectores.config(state="normal")
        self.lbl_sectores_hint.config(text=f"Sectores seleccionados: {len(sectores)} de {len(sectores)}")

        ttk.Button(self.botones_sectores, text="Todos", command=self._seleccionar_todos_sectores, style="Secondary.TButton").pack(side="left", padx=(0, 6))
        ttk.Button(self.botones_sectores, text="Ninguno", command=self._deseleccionar_todos_sectores, style="Secondary.TButton").pack(side="left", padx=(0, 6))
        ttk.Button(self.botones_sectores, text="Invertir", command=self._invertir_sectores, style="Secondary.TButton").pack(side="left")

        for i, sector in enumerate(sectores):
            var = tk.BooleanVar(value=True)
            self.sectores_vars[sector] = var
            cb = ttk.Checkbutton(
                self.frame_sectores,
                text=sector,
                variable=var,
                command=self._refrescar_lista_portales,
                style="TCheckbutton"
            )
            cb.grid(row=i, column=0, sticky="w", padx=10, pady=3)
            self.sectores_widgets[sector] = cb

    def toggle_panel_sectores(self):
        if not self.tiene_columna_sector:
            return
        if self.sector_panel_visible:
            self.panel_sectores.pack_forget()
            self.sector_panel_visible = False
            self.btn_toggle_sectores.config(text="Mostrar sectores")
        else:
            self.panel_sectores.pack(fill="x")
            self.sector_panel_visible = True
            self.btn_toggle_sectores.config(text="Ocultar sectores")
            self._filtrar_checkboxes_sectores()

    def _filtrar_checkboxes_sectores(self):
        if not self.sectores_widgets:
            return
        q = self.var_buscar_sector.get().strip().lower()
        visibles = 0
        for sector, cb in self.sectores_widgets.items():
            if not q or q in sector.lower():
                if not cb.winfo_ismapped():
                    cb.grid()
                visibles += 1
            else:
                cb.grid_remove()
        self.frame_sectores_canvas.configure(scrollregion=self.frame_sectores_canvas.bbox("all"))

    def _seleccionar_todos_sectores(self):
        for var in self.sectores_vars.values():
            var.set(True)
        self._refrescar_lista_portales()

    def _deseleccionar_todos_sectores(self):
        for var in self.sectores_vars.values():
            var.set(False)
        self._refrescar_lista_portales()

    def _invertir_sectores(self):
        for var in self.sectores_vars.values():
            var.set(not var.get())
        self._refrescar_lista_portales()

    def _actualizar_resumen_sectores(self):
        if not self.sectores_vars:
            return
        total = len(self.sectores_vars)
        activos = sum(1 for v in self.sectores_vars.values() if v.get())
        self.lbl_sectores_hint.config(text=f"Sectores seleccionados: {activos} de {total}")

    def _portales_filtrados(self):
        portales = list(self.portales_todos)
        if self.sectores_vars:
            sectores_activos = {s for s, v in self.sectores_vars.items() if v.get()}
            portales = [p for p in portales if p["sector"] in sectores_activos]

        q = getattr(self, "var_buscar_portal", tk.StringVar()).get().strip().lower()
        if q:
            portales = [
                p for p in portales
                if q in p["empresa"].lower() or q in p["sector"].lower() or q in p["enlace"].lower()
            ]
        return portales

    def _refrescar_lista_portales(self):
        for row in self.tree.get_children():
            self.tree.delete(row)
        portales = self._portales_filtrados()
        for p in portales:
            self.tree.insert("", "end", values=(p["empresa"], p["sector"], p["enlace"]))
        total = len(self.portales_todos)
        filt = len(portales)
        self.lbl_count.config(text=f"{filt} portales seleccionados" if total == filt else f"{filt} de {total} portales seleccionados")
        self._actualizar_resumen_sectores()
        self.side_portales.config(text=str(filt))
        self.btn_procesar.config(state="normal" if filt > 0 else "disabled")

    def iniciar_scraping(self):
        if self.scraping_en_curso:
            return
        portales = self._portales_filtrados()
        if not portales:
            messagebox.showwarning("Sin portales", "No hay portales seleccionados. Tildá al menos un sector.")
            return

        self.scraping_en_curso = True
        self.resultados = []
        self.controlador = ControladorEjecucion()
        self.side_ofertas.config(text="0")

        self.btn_procesar.config(state="disabled")
        self.btn_cargar.config(state="disabled")
        self.btn_pausar.config(state="normal", text="⏸  Pausar")
        self.btn_exportar_y_pausar.config(state="normal")
        self.btn_detener.config(state="normal")
        self.btn_exportar_parcial.config(state="normal")
        self._set_estado_visual("● Ejecutando", COLOR_ACENTO)
        self.progress["value"] = 0
        self.t_inicio = datetime.now()

        portales_tuplas = [(p["empresa"], p["enlace"]) for p in portales]
        thread = threading.Thread(target=self._run_scraping, args=(portales_tuplas,), daemon=True)
        thread.start()

    def toggle_pausa(self):
        if not self.controlador:
            return
        if self.controlador.pausado:
            self.controlador.reanudar()
            self.btn_pausar.config(text="⏸  Pausar")
            self._set_estado_visual("● Ejecutando", COLOR_ACENTO)
            self.log_msg("\n▶ Reanudando...")
        else:
            self.controlador.pausar()
            self.btn_pausar.config(text="▶  Reanudar")
            self._set_estado_visual("⏸ Pausado", "#f59e0b")
            self.log_msg("\n⏸ Pausando. Puede tardar unos segundos en completar las vacantes en curso...")

    def detener_scraping(self):
        if not self.controlador:
            return
        if not messagebox.askyesno("Confirmar", "¿Seguro que querés detener el scraping?\nLo que ya se haya extraído queda disponible para exportar."):
            return
        self.controlador.detener()
        self._set_estado_visual("⏹ Deteniendo", COLOR_PELIGRO)
        self.log_msg("\n⏹ Deteniendo... esperando que terminen las vacantes en curso.")

    def exportar_y_pausar(self):
        if self.controlador and self.scraping_en_curso and not self.controlador.pausado:
            self.controlador.pausar()
            self.btn_pausar.config(text="▶  Reanudar")
            self._set_estado_visual("⏸ Pausado", "#f59e0b")
            self.log_msg("\n⏸ Pausado para exportar el avance actual.")
        self.exportar_parcial()

    def exportar_parcial(self):
        if not self.resultados:
            messagebox.showinfo("Sin datos", "Todavía no hay resultados para exportar.")
            return
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        carpeta = self._carpeta_exportacion()
        carpeta.mkdir(parents=True, exist_ok=True)
        sufijo = "parcial" if self.scraping_en_curso else "final"
        filtro_tiempo = self.var_filtro_tiempo.get().lower()
        keywords = normalizar_texto(self.var_keywords.get()).replace(" ", "_")
        extra = f"_{keywords[:30]}" if keywords else ""
        ruta = carpeta / f"ofertas_{sufijo}_{filtro_tiempo}{extra}_{timestamp}.xlsx"
        try:
            guardar_excel(list(self.resultados), ruta)
            self.log_msg(f"\n💾 Exportadas {len(self.resultados)} ofertas en: {ruta.name}")
            self.archivo_salida = str(ruta)
            messagebox.showinfo("Exportado", f"Se guardaron {len(self.resultados)} ofertas en:\n{ruta}")
        except Exception as e:
            messagebox.showerror("Error al guardar", str(e))

    def _run_scraping(self, portales_tuplas):
        try:
            headless = not self.var_visible.get()
            self.log_msg("\n" + "=" * 50)
            self.log_msg(f"Iniciando scraping de {len(portales_tuplas)} portales...")
            filtro_tiempo = self.var_filtro_tiempo.get()
            filtro_keywords = self.var_keywords.get().strip()
            self.log_msg(f"Concurrencia: {CONCURRENCIA} vacantes en paralelo")
            self.log_msg(f"📅 Filtro de fecha: {filtro_tiempo}")
            self.log_msg(f"🔑 Keywords: {filtro_keywords if filtro_keywords else 'Sin filtro'}")
            self.log_msg("=" * 50)

            loop = asyncio.new_event_loop()
            self.loop_scraping = loop
            asyncio.set_event_loop(loop)
            loop.run_until_complete(scrape_portales(
                portales_tuplas,
                headless,
                self.log_msg,
                self.actualizar_progreso,
                self.controlador,
                self.resultados,
                filtro_tiempo,
                filtro_keywords
            ))
            loop.close()

            duracion = (datetime.now() - self.t_inicio).total_seconds()
            self.log_msg(f"\n⏱ Tiempo total: {duracion:.1f} segundos")
            self.log_msg(f"📊 Total de ofertas extraídas: {len(self.resultados)}")

            if self.resultados:
                timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
                carpeta = self._carpeta_exportacion()
                carpeta.mkdir(parents=True, exist_ok=True)
                filtro_tiempo = self.var_filtro_tiempo.get().lower()
                keywords = normalizar_texto(self.var_keywords.get()).replace(" ", "_")
                extra = f"_{keywords[:30]}" if keywords else ""
                ruta_salida = carpeta / f"ofertas_{filtro_tiempo}{extra}_{timestamp}.xlsx"
                guardar_excel(list(self.resultados), ruta_salida)
                self.archivo_salida = str(ruta_salida)
                self.log_msg(f"✓ Guardado automático en: {ruta_salida.name}")
                self.root.after(0, self._scraping_completado)
            else:
                self.root.after(0, lambda: messagebox.showinfo("Sin resultados", "No se extrajeron vacantes."))
        except Exception as e:
            self.log_msg(f"\n✗ Error inesperado: {e}")
            self.root.after(0, lambda: messagebox.showerror("Error", str(e)))
        finally:
            self.scraping_en_curso = False
            self.controlador = None
            self.root.after(0, self._reset_botones)

    def _reset_botones(self):
        self.btn_procesar.config(state="normal" if self._portales_filtrados() else "disabled")
        self.btn_cargar.config(state="normal")
        self.btn_pausar.config(state="disabled", text="⏸  Pausar")
        self.btn_exportar_y_pausar.config(state="disabled")
        self.btn_detener.config(state="disabled")
        self.btn_exportar_parcial.config(state="normal" if self.resultados else "disabled")
        self._set_estado_visual("✓ Finalizado", COLOR_PRIMARIO)

    def _scraping_completado(self):
        respuesta = messagebox.askyesno("¡Listo!", f"Se extrajeron {len(self.resultados)} ofertas.\nArchivo:\n{self.archivo_salida}\n\n¿Querés abrir la carpeta?")
        if respuesta:
            self.abrir_carpeta()

    def abrir_carpeta(self):
        if not self.archivo_salida:
            return
        carpeta = str(Path(self.archivo_salida).parent)
        if sys.platform == "win32":
            os.startfile(carpeta)
        elif sys.platform == "darwin":
            subprocess.Popen(["open", carpeta])
        else:
            subprocess.Popen(["xdg-open", carpeta])


def main():
    root = tk.Tk()
    App(root)
    root.mainloop()


if __name__ == "__main__":
    main()
